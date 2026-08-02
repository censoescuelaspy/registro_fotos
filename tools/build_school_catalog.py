"""Genera el catalogo publico de la muestra piloto desde la planilla oficial."""

from __future__ import annotations

import argparse
import json
import hashlib
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = (
    ROOT.parent
    / "03_DATOS"
    / "Inventarios_Escuelas"
    / "Muestra_CIALPA_Capital_Central_RUE_2026_2026-07-16.xlsx"
)
DEFAULT_SITES_SOURCE = ROOT.parents[1] / "LISTADO_ESCUELAS_CIALPA_CODIGO_AULAS_2026-07-22.xlsx"
OUTPUT = ROOT / "assets" / "data" / "pilot-schools.json"
GAS_OUTPUT = ROOT / "gas" / "SchoolsData.js"
SHEET = "muestra_piloto_def"
EXPECTED_TOTAL = 86
EXPECTED_BY_DEPARTMENT = {"CAPITAL": 15, "CENTRAL": 71}
EXPECTED_SITES = 85


def clean_text(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def coordinate(value: object, field: str, code: str) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Coordenada {field} invalida para escuela {code}") from exc
    if field == "latitud" and not -90 <= number <= 90:
        raise ValueError(f"Latitud fuera de rango para escuela {code}")
    if field == "longitud" and not -180 <= number <= 180:
        raise ValueError(f"Longitud fuera de rango para escuela {code}")
    return round(number, 7)


def school_codes(value: object) -> list[str]:
    return re.findall(r"\d+", clean_text(value))


def rue_code(value: object) -> str:
    digits = "".join(character for character in clean_text(value) if character.isdigit())
    if not digits:
        raise ValueError("Se encontro un codigo MEC/RUE vacio.")
    return digits.zfill(7)


def header_row(worksheet, required: set[str]) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        headers = [clean_text(value) for value in row]
        index = {name: position for position, name in enumerate(headers) if name}
        if required <= index.keys():
            return row_number, index
    raise ValueError(f"No se encontraron los encabezados: {', '.join(sorted(required))}")


def read_sites(source: Path) -> dict[str, dict[str, object]]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook["Edificios_85"]
    required = {"N° SITIO", "CÓDIGO(S) MEC", "AULAS EST. SITIO"}
    row_number, index = header_row(worksheet, required)
    sites_by_code: dict[str, dict[str, object]] = {}
    site_ids: set[str] = set()
    for row in worksheet.iter_rows(min_row=row_number + 1, values_only=True):
        site_value = row[index["N° SITIO"]]
        if not isinstance(site_value, (int, float)):
            continue
        site_number = int(site_value)
        site_id = f"CIALPA-S{site_number:03d}"
        raw_codes = school_codes(row[index["CÓDIGO(S) MEC"]])
        if not raw_codes:
            raise ValueError(f"El sitio {site_number} no tiene codigos MEC/RUE.")
        rue_codes = [rue_code(code) for code in raw_codes]
        site_ids.add(site_id)
        site = {
            "sitioId": site_id,
            "numeroSitio": site_number,
            "codigosSitio": raw_codes,
            "codigosRueSitio": rue_codes,
            "sedeCompartida": len(raw_codes) > 1,
            "aulasEstimadasSitio": int(row[index["AULAS EST. SITIO"]] or 0),
        }
        for code in raw_codes:
            if code in sites_by_code:
                raise ValueError(f"El codigo {code} figura en mas de un sitio fisico.")
            sites_by_code[code] = site
    if len(site_ids) != EXPECTED_SITES:
        raise ValueError(f"Se esperaban {EXPECTED_SITES} sitios y se obtuvieron {len(site_ids)}")
    return sites_by_code


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--sites-source", type=Path, default=DEFAULT_SITES_SOURCE)
    args = parser.parse_args()
    source = args.source
    sites_source = args.sites_source
    if not source.exists():
        raise FileNotFoundError(f"No se encontro la fuente de escuelas y coordenadas: {source}")
    if not sites_source.exists():
        raise FileNotFoundError(f"No se encontro la fuente de sedes fisicas: {sites_source}")

    sites_by_code = read_sites(sites_source)
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook[SHEET]
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    index = {name: position for position, name in enumerate(headers)}
    required = {
        "ENUMERA",
        "DEPTO",
        "DIST",
        "ZONA",
        "LOCALIDAD",
        "CODIGO",
        "NOMBRE",
        "LAT_DEC",
        "LNG_DEC",
    }
    missing = sorted(required - index.keys())
    if missing:
        raise ValueError(f"Columnas faltantes: {', '.join(missing)}")

    schools: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in rows:
        code = clean_text(row[index["CODIGO"]])
        if not code:
            continue
        if code in seen:
            raise ValueError(f"Codigo duplicado en la muestra: {code}")
        if code not in sites_by_code:
            raise ValueError(f"El codigo {code} no tiene sitio fisico conciliado.")
        seen.add(code)
        site = sites_by_code[code]
        schools.append(
            {
                "codigo": code,
                "codigoRue": rue_code(code),
                **site,
                "nombre": clean_text(row[index["NOMBRE"]]),
                "departamento": clean_text(row[index["DEPTO"]]).upper(),
                "distrito": clean_text(row[index["DIST"]]),
                "zona": clean_text(row[index["ZONA"]]).upper(),
                "localidad": clean_text(row[index["LOCALIDAD"]]),
                "latitud": coordinate(row[index["LAT_DEC"]], "latitud", code),
                "longitud": coordinate(row[index["LNG_DEC"]], "longitud", code),
                "ordenMuestra": int(row[index["ENUMERA"]]),
            }
        )

    schools.sort(key=lambda school: int(school["ordenMuestra"]))
    by_department = Counter(str(school["departamento"]) for school in schools)
    if len(schools) != EXPECTED_TOTAL:
        raise ValueError(f"Se esperaban {EXPECTED_TOTAL} escuelas y se obtuvieron {len(schools)}")
    if dict(by_department) != EXPECTED_BY_DEPARTMENT:
        raise ValueError(
            f"Distribucion inesperada: {dict(by_department)}; esperada: {EXPECTED_BY_DEPARTMENT}"
        )
    if set(sites_by_code) != seen:
        raise ValueError(
            "La fuente de sitios contiene codigos ausentes del catalogo: "
            + ", ".join(sorted(set(sites_by_code) - seen))
        )

    payload = {
        "schemaVersion": 2,
        "generatedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": source.name,
        "sourceSha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "sources": [
            {"name": source.name, "sha256": hashlib.sha256(source.read_bytes()).hexdigest()},
            {"name": sites_source.name, "sha256": hashlib.sha256(sites_source.read_bytes()).hexdigest()},
        ],
        "scope": "Muestra piloto Capital y Central",
        "total": len(schools),
        "physicalSites": len({str(school["sitioId"]) for school in schools}),
        "sharedSites": sum(1 for site_id in {str(school["sitioId"]) for school in schools}
                           if sum(1 for school in schools if str(school["sitioId"]) == site_id) > 1),
        "byDepartment": dict(by_department),
        "schools": schools,
    }
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    GAS_OUTPUT.write_text(
        "// Generado por tools/build_school_catalog.py. No editar manualmente.\n"
        "const PILOT_SCHOOLS = "
        + json.dumps(schools, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(f"Catalogo generado: {OUTPUT} ({len(schools)} escuelas)")
    print(f"Catalogo GAS generado: {GAS_OUTPUT}")


if __name__ == "__main__":
    main()
