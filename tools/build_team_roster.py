#!/usr/bin/env python3
"""Conciliación reproducible de equipos y escuelas para CIALPA Fotos.

La salida usa códigos operativos no personales. No extrae ni publica cédulas,
teléfonos, correos, PIN ni otros datos de los currículums.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook


PLAN_BLOCKS = (
    (1, 64, 75),
    (2, 77, 86),
    (3, 88, 96),
    (4, 98, 110),
    (5, 112, 120),
    (6, 128, 137),
    (7, 138, 148),
    (8, 149, 160),
)

NAME_OVERRIDES = {
    "COLEGIO NACIONAL SAGRADA FAMILIA": 73,
    "COLEGIO NACIONAL DR EMILIO CUBAS": 71,
    "COLEGIO NACIONAL HEROES DEL CHACO": 75,
    "COLEGIO NACIONAL SAN AGUSTIN": 76,
    "COLEGIO NACIONAL SANTA LUCIA": 78,
    "COLEGIO NACIONAL TACIANA DE VILLALBA": 74,
    "COLEGIO NACIONAL SAN JOSE": 60,
    "ESCUELA BASICA EPIFANIO MENDEZ FLEITAS JULIO CORREA": 42,
}


def text(value: object) -> str:
    return "" if value is None else " ".join(str(value).strip().split())


def normalize(value: object) -> str:
    value = unicodedata.normalize("NFD", text(value))
    value = value.encode("ascii", "ignore").decode().upper()
    value = value.replace("M.R.A.", "MARIANO ROQUE ALONSO")
    return re.sub(r"[^A-Z0-9]+", " ", value).strip()


def codes(value: object) -> list[str]:
    return re.findall(r"\d+", text(value))


def match_score(plan: dict, site: dict) -> float:
    left = normalize(plan["name"])
    right = normalize(site["name"])
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    sequence = SequenceMatcher(None, left, right).ratio()
    jaccard = len(left_tokens & right_tokens) / max(1, len(left_tokens | right_tokens))
    district = 0.10 if normalize(plan["district"]) == normalize(site["district"]) else 0
    return 0.55 * sequence + 0.35 * jaccard + district


def read_plan(base: Path) -> tuple[list[dict], dict[int, list[str]]]:
    source = base / "R04_PLAN OPERATIVO PLAN PILOTO_EQUIPOS Y CRONOGRAMA (1).xlsx"
    sheet = load_workbook(source, read_only=True, data_only=True)["Hoja1"]
    rows: list[dict] = []
    teams: dict[int, list[str]] = {}
    for team, first, last in PLAN_BLOCKS:
        members = [item.strip() for item in str(sheet.cell(first, 38).value or "").splitlines()]
        if len(members) != 2:
            raise ValueError(f"Equipo {team}: se esperaban dos integrantes.")
        teams[team] = members
        for row in range(first, last + 1):
            name = text(sheet.cell(row, 40).value)
            if not name:
                continue
            rows.append(
                {
                    "team": team,
                    "name": name,
                    "district": text(sheet.cell(row, 42).value),
                    "visit_order": int(sheet.cell(row, 39).value),
                    "source_row": row,
                }
            )
    if len(rows) != 85:
        raise ValueError(f"El plan debe contener 85 sitios; contiene {len(rows)}.")
    return rows, teams


def read_sites(base: Path) -> list[dict]:
    source = base / "LISTADO_ESCUELAS_CIALPA_CODIGO_AULAS_2026-07-22.xlsx"
    sheet = load_workbook(source, read_only=True, data_only=True)["Edificios_85"]
    rows: list[dict] = []
    for row in range(5, sheet.max_row + 1):
        number = sheet.cell(row, 1).value
        if not isinstance(number, (int, float)):
            continue
        rows.append(
            {
                "site": int(number),
                "codes": codes(sheet.cell(row, 3).value),
                "name": text(sheet.cell(row, 4).value),
                "district": text(sheet.cell(row, 6).value),
            }
        )
    if len(rows) != 85:
        raise ValueError(f"El listado debe contener 85 sitios; contiene {len(rows)}.")
    return rows


def match_sites(plan: list[dict], sites: list[dict]) -> list[dict]:
    site_by_number = {site["site"]: site for site in sites}
    used_plan: set[int] = set()
    used_sites: set[int] = set()
    matches: list[dict] = []
    for item in plan:
        site_number = NAME_OVERRIDES.get(normalize(item["name"]))
        if not site_number:
            continue
        matches.append({"plan": item, "site": site_by_number[site_number], "score": 1.0})
        used_plan.add(item["source_row"])
        used_sites.add(site_number)

    candidates = sorted(
        (
            (match_score(item, site), item, site)
            for item in plan
            if item["source_row"] not in used_plan
            for site in sites
            if site["site"] not in used_sites
        ),
        reverse=True,
        key=lambda candidate: candidate[0],
    )
    for score, item, site in candidates:
        if item["source_row"] in used_plan or site["site"] in used_sites:
            continue
        matches.append({"plan": item, "site": site, "score": score})
        used_plan.add(item["source_row"])
        used_sites.add(site["site"])

    if len(matches) != 85:
        raise ValueError(f"Conciliación incompleta: {len(matches)} de 85 sitios.")
    weak = [item for item in matches if item["score"] < 0.50]
    if weak:
        raise ValueError(
            "Conciliaciones débiles: "
            + ", ".join(
                f"fila {item['plan']['source_row']} / sitio {item['site']['site']}"
                for item in weak
            )
        )
    return matches


def split_name(full_name: str) -> tuple[str, str]:
    parts = full_name.split()
    if len(parts) < 2:
        raise ValueError(f"Nombre incompleto: {full_name!r}")
    return parts[0], " ".join(parts[1:])


def build_payload(base: Path, catalog_path: Path) -> dict:
    plan, teams = read_plan(base)
    sites = read_sites(base)
    matches = match_sites(plan, sites)
    catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    catalog_codes = {str(item["codigo"]) for item in catalog["schools"]}

    users: list[dict] = []
    representatives: dict[int, str] = {}
    for team, members in sorted(teams.items()):
        for member_index, full_name in enumerate(members, start=1):
            code = f"260{team:02d}{member_index}"
            names, surnames = split_name(full_name)
            if member_index == 1:
                representatives[team] = code
            users.append(
                {
                    "codigo_censista": code,
                    "nombres": names,
                    "apellidos": surnames,
                    "rol": "ENCUESTADOR",
                    "pin_salt": "",
                    "pin_hash": "",
                    "activo": True,
                    "telefono": "",
                    "created_at": "2026-07-25T00:00:00-03:00",
                    "updated_at": "2026-07-25T00:00:00-03:00",
                    "ultimo_acceso": "",
                    "equipo": f"Equipo {team}",
                }
            )

    assignments: list[dict] = []
    assigned_codes: set[str] = set()
    team_site_counts: dict[int, int] = {team: 0 for team in teams}
    for item in matches:
        team = item["plan"]["team"]
        site = item["site"]
        members = teams[team]
        team_site_counts[team] += 1
        for school_code in site["codes"]:
            assigned_codes.add(school_code)
            assignments.append(
                {
                    "assignment_id": f"CIALPA-FOTOS-2026-E{team:02d}-{school_code}",
                    "codigo_censista": representatives[team],
                    "codigo_escuela": school_code,
                    "activo": True,
                    "fecha_asignacion": "2026-07-25T00:00:00-03:00",
                    "asignado_por": "admin",
                    "notas": (
                        f"Equipo {team} — {members[0]} / {members[1]} | "
                        f"Sitio físico {site['site']} | Orden de visita {item['plan']['visit_order']}"
                    ),
                    "updated_at": "2026-07-25T00:00:00-03:00",
                }
            )

    if assigned_codes != catalog_codes:
        raise ValueError(
            f"Códigos no conciliados. Faltan={sorted(catalog_codes-assigned_codes)}; "
            f"sobran={sorted(assigned_codes-catalog_codes)}"
        )
    return {
        "summary": {
            "teams": len(teams),
            "users": len(users),
            "sites": len(sites),
            "school_codes": len(assignments),
            "team_site_counts": team_site_counts,
        },
        "users": users,
        "assignments": assignments,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", type=Path, default=Path(r"J:\Mi unidad"))
    parser.add_argument(
        "--catalog",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "assets" / "data" / "pilot-schools.json",
    )
    parser.add_argument("--summary", action="store_true")
    args = parser.parse_args()
    payload = build_payload(args.base, args.catalog)
    print(
        json.dumps(
            payload["summary"] if args.summary else payload,
            ensure_ascii=False,
            separators=(",", ":"),
        )
    )


if __name__ == "__main__":
    main()
